# -*- coding: utf-8 -*-
"""
Created on Thu Aug 12 21:16:31 2021

@author: JineshJhonsa
"""
import openpyxl
import numpy as np

data = np.zeros([100, 2], dtype = float)


def add_data(x,vdd):
    global ind,data,toffset
    ind = ind +1
    test = '{0:.12f}'.format(x*toffset)
    data[ind] =[test,vdd]
    
def add_coeff(coe,x):
    global toffset
    add_data(x,vdd)
    if coe == 1:
         wbll.write('%s %s \\\n' % ('{0:.12f}'.format((x)*toffset), '{0:.12f}'.format(vdd)))
         wblr.write('%s %s \\\n' % ('{0:.12f}'.format((x)*toffset), '{0:.12f}'.format(0)))
    if coe == -1:
         wbll.write('%s %s \\\n' % ('{0:.12f}'.format((x)*toffset), '{0:.12f}'.format(0)))
         wblr.write('%s %s \\\n' % ('{0:.12f}'.format((x)*toffset), '{0:.12f}'.format(vdd)))
    if coe == 0:
         wbll.write('%s %s \\\n' % ('{0:.12f}'.format((x)*toffset), '{0:.12f}'.format(0)))
         wblr.write('%s %s \\\n' % ('{0:.12f}'.format((x)*toffset), '{0:.12f}'.format(0)))

wbll = open('write_bitline_left_coeff.txt', 'w')
wblr =  open('write_bitline_right_coeff.txt', 'w')
wwl  = open('write_word_coeff.txt', 'w')
wwl.write('simulator lang=spectre\n')
wbll.write('simulator lang=spectre\n')
wblr.write('simulator lang=spectre\n')


wwl.write('va3 (a\\<3\\> VSS) vsource dc=0 \n')
wwl.write('va2 (a\\<2\\> VSS) vsource dc=0.8 \n')
wwl.write('va1 (a\\<1\\> VSS) vsource dc=0.8 \n')
wwl.write('va0 (a\\<0\\> VSS) vsource dc=0.8 \n')


wb = openpyxl.load_workbook("coeff_file.xlsx")
ws = wb['sram']
coeff =  wb['coeff']
ind =0
i = 0
vdd=0.8
coe =[]
for r in coeff.iter_rows():
    x = r[0].value
    coe.append(x)

for row in ws.iter_rows():
    num = str(i)
    x = row[0].value
    if x == 1:
        wwl.write('vsram_'+ num +' (INIT'+num+' VSS) vsource dc=0.8 \n')
    else:
        wwl.write('vsram_'+ num +' (INIT'+num+' VSS) vsource dc=0 \n')
    i =i +1
    
toffset = 0.000000001
tr = 0.00000000002
offval =0
pwidth =5
# initializing pwl wave for wwl wbl wbr

wwl.write('\n')
wbll.write('vwbl_0 (WBLL VSS) vsource type=pwl wave=[  \\\n')
wblr.write('vwrl_0 (WBLR VSS) vsource type=pwl wave=[  \\\n')  
wwl.write('vwwl_%s (WWL%s VSS) vsource type=pwl wave=[  \\\n'%(0, 0))  

#  Adding  coefficient 0  to wwl wbl wbr
for i in range(pwidth):    
    x = i +offval
    wwl.write('%s %s \\\n' % ('{0:.12f}'.format((x)*toffset), '{0:.12f}'.format(vdd)))
    if i ==0:
        add_coeff(coe[0],x+0.02)
    else:
        add_coeff(coe[0], x)

wwl.write('%s %s] \n' % ('{0:.12f}'.format((x+0.02)*toffset), '{0:.12f}'.format(0)))   
# adding one extra coefficient 0 for 10 ns
add_coeff(coe[0], x+1)
# falling to zero 
add_coeff(0, x+1.02)
# addingn two zeros to the wbl wbr
add_coeff(0, x+2)
add_coeff(0, x+3)

# updateing offval for next write pulse
offval = offval+pwidth+2  
wwl.write('\n')

wwl.write('vwwl_%s (WWL%s VSS) vsource type=pwl wave=[  \\\n'%(1, 1))  

for i in range (offval):
    wwl.write('%s %s \\\n' % ('{0:.12f}'.format((i)*toffset), '{0:.12f}'.format(0)))
wwl.write('%s %s \\\n' % ('{0:.12f}'.format((i+0.02)*toffset), '{0:.12f}'.format(vdd)))   
for i in range(pwidth):
    x = i +offval
    wwl.write('%s %s \\\n' % ('{0:.12f}'.format((x)*toffset), '{0:.12f}'.format(vdd)))
    if i ==0:
        add_coeff(coe[1],x+0.02)
    else:
        add_coeff(coe[1], x)
        
wwl.write('%s %s] \n' % ('{0:.12f}'.format((x+0.02)*toffset), '{0:.12f}'.format(0)))    

add_coeff(coe[1], x+1)
add_coeff(0, x+1.02) 
add_coeff(0, x+2)
add_coeff(0, x+3)   

offval = pwidth + offval+2   

wwl.write('\n')

wwl.write('vwwl_%s (WWL%s VSS) vsource type=pwl wave=[  \\\n'%(2, 2))  

for i in range (offval):
    wwl.write('%s %s \\\n' % ('{0:.12f}'.format((i)*toffset), '{0:.12f}'.format(0)))
wwl.write('%s %s \\\n' % ('{0:.12f}'.format((i+0.02)*toffset), '{0:.12f}'.format(vdd)))
for i in range(pwidth):
    x = i +offval
    wwl.write('%s %s \\\n' % ('{0:.12f}'.format((x)*toffset), '{0:.12f}'.format(vdd)))
    if i ==0:
        add_coeff(coe[2],x+0.02)
    else:
        add_coeff(coe[2], x)
wwl.write('%s %s \\\n' % ('{0:.12f}'.format((x+0.02)*toffset), '{0:.12f}'.format(0)))
    
add_coeff(coe[2], x+1)
add_coeff(0, x+1.002)   
add_coeff(0, x+2)
add_coeff(0, x+3)    
wwl.write('%s %s] \n' % ('{0:.12f}'.format((x+1)*toffset), '{0:.12f}'.format(0)))    
offval = pwidth + offval+2 

wwl.write('\n')

wwl.write('vwwl_%s (WWL%s VSS) vsource type=pwl wave=[  \\\n'%(3, 3))  

for i in range (offval):
    wwl.write('%s %s \\\n' % ('{0:.12f}'.format((i)*toffset), '{0:.12f}'.format(0)))
wwl.write('%s %s \\\n' % ('{0:.12f}'.format((i+0.02)*toffset), '{0:.12f}'.format(vdd)))

for i in range(pwidth):
    x = i +offval
    wwl.write('%s %s \\\n' % ('{0:.12f}'.format((x)*toffset), '{0:.12f}'.format(vdd)))
    if i ==0:
        add_coeff(coe[3],x+0.02)
    else:
        add_coeff(coe[3], x)
wwl.write('%s %s \\\n' % ('{0:.12f}'.format((x+0.02)*toffset), '{0:.12f}'.format(0)))

add_coeff(coe[3], x+1)
add_coeff(0, x+1.002)    
add_coeff(0, x+2)
add_coeff(0, x+3)
wwl.write('%s %s] \n' % ('{0:.12f}'.format((x+1)*toffset), '{0:.12f}'.format(0)))    
offval = pwidth + offval + 2

wbll.write('%s %s] \n' % ('{0:.12f}'.format((x+4)*toffset), '{0:.12f}'.format(0)))   
wblr.write('%s %s] \n' % ('{0:.12f}'.format((x+4)*toffset), '{0:.12f}'.format(0)))   
 
wwl.close()
wbll.close()
wblr.close()

pulse  = open('enable.txt', 'w')
pulse.write('simulator lang=spectre\n')


pulse.write('vclk (CLK VSS) vsource type=pulse val0=0 val1=0.8 delay=25n rise= 0.02n fall= 0.02n width=5n period=10n')
pulse.write('\n')

pulse.write('vs_en (EN VSS) vsource type=pulse val0=0 val1=0.8 delay=30n rise= 0.02n fall= 0.02n width=10n period=20n')
pulse.write('\n')

pulse.write('vspc (SPC VSS) vsource type=pulse val0=0 val1=0.8 delay=30n rise= 0.02n fall= 0.02n width=10n period=20n')
pulse.write('\n')
 # one time mux en
pulse.write('vmux_en (MUX_EN VSS) vsource type=pulse val0=0 val1=0.8 delay=40n rise= 0.02n fall= 0.02n width=9u period=10u')
pulse.write('\n')
pulse.write('vpc (PC VSS) vsource type=pulse val0=0 val1=0.8 delay=30n rise= 0.02n fall= 0.02n width=15n period=20n')
pulse.write('\n')
pulse.write('vsrl (SWL VSS) vsource type=pulse val0=0 val1=0.8 delay=30n rise= 0.02n fall= 0.02n width=10n period=20n')
pulse.write('\n')
pulse.write('vsa_en (SA_EN VSS) vsource type=pulse val0=0 val1=0.8 delay=40n rise= 0.02n fall= 0.02n width=10n period=20n')
pulse.write('\n')

pulse.close()